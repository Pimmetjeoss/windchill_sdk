"""Tests for the SAP BOM export reader."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from windchill_spm.sap_reader import _cell_int, _cell_str, _normalize_wh, read_sap_export


class TestNormalizeWh:
    def test_removes_dot(self):
        assert _normalize_wh("WH.123456") == "WH123456"

    def test_no_dot(self):
        assert _normalize_wh("WH123456") == "WH123456"

    def test_empty(self):
        assert _normalize_wh("") == ""

    def test_none(self):
        assert _normalize_wh(None) == ""

    def test_non_wh(self):
        assert _normalize_wh("ABC123") == "ABC123"


class TestCellConversions:
    def test_cell_str_none(self):
        assert _cell_str(None) == ""

    def test_cell_str_float_int(self):
        assert _cell_str(2.0) == "2"

    def test_cell_str_float_decimal(self):
        assert _cell_str(2.5) == "2.5"

    def test_cell_str_string(self):
        assert _cell_str("hello") == "hello"

    def test_cell_int_none(self):
        assert _cell_int(None) == 0

    def test_cell_int_string(self):
        assert _cell_int("3") == 3

    def test_cell_int_float(self):
        assert _cell_int(2.0) == 2

    def test_cell_int_invalid(self):
        assert _cell_int("abc") == 0


# 16-column SAP EXPORT.XLSX headers
_HEADERS = [
    "Explosion level", "Object description", "Component number", "Document",
    "Item Number", "Item category", "Sort String", "Component number",
    "Object description", "Item Text", "Document Sort String",
    "Special procurement", "Document Type", "Document",
    "Document Version", "Document Part",
]


class TestReadSapExport:
    def _create_test_xlsx(self, rows: list[list]) -> Path:
        """Create a temporary xlsx file with the given rows (16 columns)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, val in enumerate(_HEADERS, 1):
            ws.cell(row=1, column=col, value=val)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        path = Path(tempfile.mktemp(suffix=".xlsx"))
        wb.save(path)
        wb.close()
        return path

    def test_reads_16col_export(self):
        path = self._create_test_xlsx([
            # level, parent_name, parent_no, document, item_no, item_cat,
            # sort_string, component_no, obj_desc, item_text,
            # doc_sort_string, special_proc, doc_type, parent_dwg_no,
            # doc_version, doc_part
            [1, "", "", "WH1065253", "0010", "L", "", "WHD23001",
             "ASSEMBLY-ORDER 1", "", "", "", "", "", "", ""],
            [2, "", "", "WH1065253", "0504", "L", "0200", "WH1066715-B00",
             "LIFT TABLE", "", "000", "50", "ZAS", "WH1066715", "00", "000"],
        ])
        try:
            rows = read_sap_export(path)
            assert len(rows) == 2
            assert rows[0].level == 1
            assert rows[0].component_number == "WHD23001"
            assert rows[0].document == "WH1065253"
            assert rows[0].main_overview_drawing == "WH1065253"  # Property alias
            assert rows[1].component_number == "WH1066715-B00"
            assert rows[1].sort_string == "0200"
            assert rows[1].tsl_number == "0200"  # Property alias
            assert rows[1].document_type == "ZAS"
            assert rows[1].parent_drawing_number == "WH1066715"
            assert rows[1].special_procurement == "50"
        finally:
            path.unlink(missing_ok=True)

    def test_normalizes_wh_dots(self):
        path = self._create_test_xlsx([
            [2, "", "WH.100000", "", "10", "L", "100", "WH.500000",
             "Test", "", "", "", "", "", "", ""],
        ])
        try:
            rows = read_sap_export(path)
            assert rows[0].parent_number == "WH100000"
            assert rows[0].component_number == "WH500000"
        finally:
            path.unlink(missing_ok=True)

    def test_skips_empty_rows(self):
        path = self._create_test_xlsx([
            [2, "", "WH100000", "", "10", "L", "100", "WH300000",
             "Part A", "", "", "", "", "", "", ""],
            [None] * 16,
        ])
        try:
            rows = read_sap_export(path)
            assert len(rows) == 1
        finally:
            path.unlink(missing_ok=True)

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            read_sap_export("/nonexistent/file.xlsx")

    def test_empty_file(self):
        wb = openpyxl.Workbook()
        path = Path(tempfile.mktemp(suffix=".xlsx"))
        wb.save(path)
        wb.close()
        try:
            with pytest.raises(ValueError, match="empty"):
                read_sap_export(path)
        finally:
            path.unlink(missing_ok=True)

    def test_pads_short_rows(self):
        """Rows with fewer than 16 columns should be padded."""
        path = self._create_test_xlsx([
            [2, "", "WH100000", "", "10", "L", "100", "WH300000",
             "Part A"],  # Only 9 columns
        ])
        try:
            rows = read_sap_export(path)
            assert len(rows) == 1
            assert rows[0].document_type == ""  # Padded column
        finally:
            path.unlink(missing_ok=True)
